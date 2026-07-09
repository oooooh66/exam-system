"""
成绩报表模块 - API 视图

提供以下接口：
- GET  /api/reports/exam/{id}/statistics/  考试统计（教师/管理员）
- GET  /api/reports/student-scores/        学生个人成绩列表
- GET  /api/reports/exam/{id}/export/      导出成绩 Excel（管理员）
"""
import io
from django.db import models as django_models
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from apps.exams.models import BusiExamSession, BusiExamSubmission, BusiStudentAnswer
from apps.exams.serializers import BusiExamSubmissionSerializer, BusiStudentAnswerSerializer
from utils.permissions import IsAdmin, IsTeacher
from utils.response import APIResponse


class BusiExamStatisticsView(APIView):
    """
    考试统计接口

    GET /api/reports/exam/{exam_id}/statistics/
    权限：教师和管理员
    返回：参考人数、平均分、最高分、最低分、及格率、各分数段分布
    """
    permission_classes = [IsTeacher]

    def get(self, request, exam_id):
        try:
            exam = BusiExamSession.objects.get(id=exam_id, is_deleted=False)
        except BusiExamSession.DoesNotExist:
            return APIResponse.error(code=404, message='考试场次不存在')

        submissions = BusiExamSubmission.objects.filter(
            exam_session=exam,
            status__in=('submitted', 'auto_submitted'),
        ).select_related('student')

        total = submissions.count()
        if total == 0:
            return APIResponse.success(data={
                'total_students': 0,
                'message': '还没有学生提交考试',
            })

        # 统计分数
        scores = [s.total_score or 0 for s in submissions]
        avg_score = round(sum(scores) / len(scores), 1)
        max_score = max(scores)
        min_score = min(scores)

        # 及格率
        pass_score = exam.paper.pass_score
        passed = sum(1 for s in scores if s >= pass_score)
        pass_rate = round(passed / total * 100, 1)

        # 分数段分布
        dist = {
            '0-59': sum(1 for s in scores if s < 60),
            '60-69': sum(1 for s in scores if 60 <= s < 70),
            '70-79': sum(1 for s in scores if 70 <= s < 80),
            '80-89': sum(1 for s in scores if 80 <= s < 90),
            '90-100': sum(1 for s in scores if s >= 90),
        }

        # 学生成绩列表
        student_scores = BusiExamSubmissionSerializer(submissions, many=True).data

        return APIResponse.success(data={
            'exam_id': exam.id,
            'exam_name': exam.name,
            'paper_total_score': exam.paper.total_score,
            'pass_score': pass_score,
            'total_students': total,
            'average_score': avg_score,
            'max_score': max_score,
            'min_score': min_score,
            'pass_rate': pass_rate,
            'score_distribution': dist,
            'student_scores': student_scores,
        })


class BusiStudentScoresView(APIView):
    """学生个人成绩列表"""
    permission_classes = []

    def get(self, request):
        user = request.user
        if not user.is_authenticated:
            return APIResponse.error(code=401, message='请先登录')

        submissions = BusiExamSubmission.objects.filter(
            student=user,
            status__in=('submitted', 'auto_submitted'),
        ).select_related('exam_session', 'exam_session__paper').order_by('-submit_time')

        data = []
        for sub in submissions:
            data.append({
                'submission_id': sub.id,
                'exam_id': sub.exam_session.id,
                'exam_name': sub.exam_session.name,
                'paper_name': sub.exam_session.paper.name,
                'total_score': sub.exam_session.paper.total_score,
                'score_obtained': sub.total_score,
                'status': sub.status,
                'submit_time': sub.submit_time,
            })

        return APIResponse.success(data=data)


class BusiExamResultDetailView(APIView):
    """
    某场考试的详细结果（学生查看自己的答题详情）
    """
    permission_classes = []

    def get(self, request, submission_id):
        try:
            submission = BusiExamSubmission.objects.get(
                id=submission_id,
                student=request.user,
                status__in=('submitted', 'auto_submitted'),
            )
        except BusiExamSubmission.DoesNotExist:
            return APIResponse.error(code=404, message='提交记录不存在')

        answers = BusiStudentAnswer.objects.filter(
            exam_session=submission.exam_session,
            student=request.user,
        ).select_related('paper_question__question').order_by('paper_question__order')

        return APIResponse.success(data={
            'submission': BusiExamSubmissionSerializer(submission).data,
            'answers': BusiStudentAnswerSerializer(answers, many=True).data,
        })


class BusiExportScoresView(APIView):
    """
    导出考试成绩 Excel

    GET /api/reports/exam/{exam_id}/export/
    权限：管理员
    返回：Excel 文件流
    """
    permission_classes = [IsAdmin]

    def get(self, request, exam_id):
        try:
            exam = BusiExamSession.objects.get(id=exam_id, is_deleted=False)
        except BusiExamSession.DoesNotExist:
            return APIResponse.error(code=404, message='考试场次不存在')

        submissions = BusiExamSubmission.objects.filter(
            exam_session=exam,
            status__in=('submitted', 'auto_submitted'),
        ).select_related('student')

        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{exam.name}成绩单'

        # 样式定义
        header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
        header_fill = openpyxl.styles.PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # 表头
        headers = ['序号', '学号', '姓名', '总分', '得分', '正确率', '提交时间', '考试开始', '考试结束']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 数据行
        for idx, sub in enumerate(submissions, 1):
            total = exam.paper.total_score or 1
            accuracy = f'{round((sub.total_score or 0) / total * 100, 1)}%' if total > 0 else '0%'
            row_data = [
                idx,
                sub.student.username,
                sub.student.username,
                exam.paper.total_score,
                sub.total_score or 0,
                accuracy,
                sub.submit_time.strftime('%Y-%m-%d %H:%M') if sub.submit_time else '',
                exam.start_time.strftime('%Y-%m-%d %H:%M'),
                exam.end_time.strftime('%Y-%m-%d %H:%M'),
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 自动调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件
        from django.http import HttpResponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{exam.name}_成绩单.xlsx"'
        return response
