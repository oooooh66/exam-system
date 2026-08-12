"""
考试模块 - API 视图

提供以下接口：
- 教师/管理员：发布考试、查看考试列表
- 学生：查看可参加的考试、开始答题、提交答案、暂存答案
"""
from django.utils import timezone
from django.db import transaction, models
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.exams.models import BusiExamSession, BusiStudentAnswer, BusiExamSubmission
from apps.users.models import BusiUser
from apps.exams.serializers import (
    BusiExamSessionListSerializer,
    BusiExamSessionDetailSerializer,
    BusiExamSessionCreateSerializer,
    BusiAnswerSubmitSerializer,
    BusiStudentAnswerSerializer,
    BusiExamSubmissionSerializer,
)
from apps.exams.scoring import auto_grade_question, grade_exam_submission
from apps.papers.models import BusiPaperQuestion
from utils.permissions import IsTeacher, IsAdmin
from utils.response import APIResponse


class BusiExamSessionViewSet(viewsets.ModelViewSet):
    """考试场次管理 ViewSet"""
    queryset = BusiExamSession.objects.filter(is_deleted=False).select_related('paper')
    permission_classes = [IsTeacher]

    def get_serializer_class(self):
        if self.action == 'create':
            return BusiExamSessionCreateSerializer
        if self.action in ('list',):
            return BusiExamSessionListSerializer
        return BusiExamSessionDetailSerializer

    def get_permissions(self):
        """学生也可以查看可用考试列表"""
        if self.action in ('list', 'retrieve', 'start_exam', 'save_answer',
                           'submit_exam', 'my_exams', 'my_result'):
            return []
        return [IsTeacher()]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.save(update_fields=['is_deleted', 'updated_at'])

    def get_queryset(self):
        qs = super().get_queryset()
        # 学生只看分配给自己的或开放的考试
        if self.request.user.is_student:
            qs = qs.filter(
                models.Q(students__isnull=True) | models.Q(students=self.request.user)
            ).distinct()
        return qs

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        return APIResponse.success(data=response.data)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = BusiExamSessionDetailSerializer(instance, context={'request': request})
        return APIResponse.success(data=serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = BusiExamSessionCreateSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return APIResponse.error(code=400, message=self._first_error(serializer.errors))
        exam = serializer.save()
        return APIResponse.created(
            data=BusiExamSessionListSerializer(exam).data,
            message='考试发布成功',
        )

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return APIResponse.success(message='考试已删除')

    @action(methods=['post'], detail=True, url_path='start')
    def start_exam(self, request, pk=None):
        """学生开始答题（动态抽题）"""
        import random
        from apps.exams.models import ExamRule
        from apps.questions.models import BusiQuestion

        exam = self.get_object()
        now = timezone.now()

        if now < exam.start_time:
            return APIResponse.error(code=400, message='考试尚未开始')
        if now > exam.end_time:
            return APIResponse.error(code=400, message='考试已结束')

        submission, created = BusiExamSubmission.objects.get_or_create(
            exam_session=exam, student=request.user,
            defaults={'status': 'in_progress', 'start_time': now},
        )
        if submission.status in ('submitted', 'auto_submitted'):
            return APIResponse.error(code=400, message='您已提交过本场考试')

        if created:
            rules = ExamRule.objects.filter(exam_session=exam).order_by('question_type')
            if not rules.exists():
                return APIResponse.error(code=400, message='该考试未配置抽题规则')

            from apps.questions.models import BusiQuestionCategory
            org_no = getattr(request.user, 'org_no', '')
            org_prefix = org_no[:3] if org_no else ''

            drawn = []
            order = 0
            for rule in rules:
                qs = BusiQuestion.objects.filter(
                    is_deleted=False, question_type=rule.question_type,
                )
                cat_filter_applied = False
                all_ids: list[int] = []

                if rule.categories:
                    # 区分指标分类和普通分类
                    cat_names = dict(BusiQuestionCategory.objects.filter(
                        id__in=rule.categories,
                    ).values_list('id', 'name'))
                    indicator_ids = [cid for cid, name in cat_names.items() if '指标' in name]
                    regular_ids = [cid for cid, name in cat_names.items() if '指标' not in name]

                    if regular_ids:
                        cat_filter_applied = True
                        all_ids.extend(list(
                            qs.filter(category_id__in=regular_ids).values_list('id', flat=True)
                        ))
                    if indicator_ids and org_prefix:
                        cat_filter_applied = True
                        all_ids.extend(list(
                            qs.filter(category_id__in=indicator_ids).filter(
                                org_id__startswith=org_prefix,
                                data_dt=rule.data_dt,
                            ).values_list('id', flat=True)
                        ))

                if not cat_filter_applied:
                    all_ids = list(qs.values_list('id', flat=True))
                sampled = random.sample(all_ids, min(rule.count, len(all_ids)))
                for qid in sampled:
                    question = BusiQuestion.objects.get(id=qid)
                    pq, _ = BusiPaperQuestion.objects.get_or_create(
                        paper=exam.paper, question=question,
                        defaults={'order': order, 'score': question.default_score},
                    )
                    BusiStudentAnswer.objects.get_or_create(
                        exam_session=exam, student=request.user, paper_question=pq,
                        defaults={'answer': None, 'status': 'draft'},
                    )
                    drawn.append(qid)
                    order += 1

            submission.drawn_question_ids = drawn
            submission.save(update_fields=['drawn_question_ids'])

        # 返回试卷信息（已有答题记录或刚创建的）
        pqs = BusiPaperQuestion.objects.filter(
            question_id__in=submission.drawn_question_ids,
            paper=exam.paper,
        ).select_related('question').order_by('order')

        paper_data = {
            'exam_id': exam.id, 'exam_name': exam.name,
            'paper_id': exam.paper.id, 'paper_name': exam.paper.name,
            'duration_minutes': exam.paper.duration_minutes,
            'total_score': sum(pq.score for pq in pqs),
            'start_time': exam.start_time, 'end_time': exam.end_time,
            'submission_start_time': submission.start_time,
            'questions': [],
        }
        for pq in pqs:
            answer = BusiStudentAnswer.objects.filter(
                exam_session=exam, student=request.user, paper_question=pq,
            ).first()
            paper_data['questions'].append({
                'paper_question_id': pq.id, 'order': pq.order, 'score': pq.score,
                'question_type': pq.question.question_type,
                'content': pq.question.content, 'options': pq.question.options,
                'saved_answer': answer.answer if answer else None,
                'status': answer.status if answer else 'draft',
            })

        return APIResponse.success(data=paper_data, message='考试开始')

    @action(methods=['post'], detail=True, url_path='save')
    def save_answer(self, request, pk=None):
        """暂存答案（不提交）"""
        exam = self.get_object()
        now = timezone.now()

        if now > exam.end_time:
            return APIResponse.error(code=400, message='考试已结束，无法保存')

        paper_question_id = request.data.get('paper_question_id')
        answer_data = request.data.get('answer')

        if not paper_question_id:
            return APIResponse.error(code=400, message='缺少题目 ID')

        try:
            answer = BusiStudentAnswer.objects.get(
                exam_session=exam,
                student=request.user,
                paper_question_id=paper_question_id,
            )
        except BusiStudentAnswer.DoesNotExist:
            return APIResponse.error(code=404, message='无效的题目')

        answer.answer = answer_data
        answer.status = 'draft'
        answer.save(update_fields=['answer', 'status', 'updated_at'])

        return APIResponse.success(message='答案已保存')

    @action(methods=['post'], detail=True, url_path='submit')
    def submit_exam(self, request, pk=None):
        """学生提交考试"""
        exam = self.get_object()
        now = timezone.now()

        if now > exam.end_time:
            # 超时自动提交
            pass
        elif now < exam.start_time:
            return APIResponse.error(code=400, message='考试尚未开始')

        # 获取提交记录
        try:
            submission = BusiExamSubmission.objects.get(
                exam_session=exam,
                student=request.user,
            )
        except BusiExamSubmission.DoesNotExist:
            return APIResponse.error(code=400, message='请先开始考试')

        if submission.status == 'submitted':
            return APIResponse.error(code=400, message='您已提交过本场考试')

        # 检查是否所有题目都已作答（考试时间到自动提交时允许跳过）
        if not request.data.get('force'):
            answers = BusiStudentAnswer.objects.filter(
                exam_session=exam, student=request.user,
            )
            total = answers.count()
            empty = 0
            for a in answers:
                ans = a.answer
                if ans is None or ans == '' or ans == []:
                    empty += 1
            if total == 0:
                return APIResponse.error(code=400, message='未检测到答题记录')
            if empty > 0:
                return APIResponse.error(code=400, message=f'还有 {empty} 道题未作答，请全部完成后再提交')

        # 批量更新答案状态
        with transaction.atomic():
            BusiStudentAnswer.objects.filter(
                exam_session=exam,
                student=request.user,
            ).update(status='submitted')

            # 自动批改客观题并计算总分
            grade_exam_submission(submission)

            submission.status = 'submitted'
            submission.submit_time = now
            submission.save()

        return APIResponse.success(
            data={
                'total_score': submission.total_score,
                'submit_time': submission.submit_time,
            },
            message='提交成功',
        )

    @action(methods=['get'], detail=False, url_path='my-exams')
    def my_exams(self, request):
        """学生查看我的考试列表"""
        if not request.user.is_student:
            qs = self.get_queryset().filter(created_by=request.user)
        else:
            qs = self.get_queryset().filter(
                models.Q(students__isnull=True) | models.Q(students=request.user)
            ).distinct()

        serializer = BusiExamSessionListSerializer(qs, many=True)
        data = list(serializer.data)

        # 为学生过滤：按业务范围（business_scope）匹配考试分类前缀
        if request.user.is_student:
            student_scope = set(getattr(request.user, 'business_scope', []))
            qs_with_scope = qs.all()  # re-fetch to get model instances for exam_scope
            scope_map = {e.id: set(e.exam_scope or []) for e in qs_with_scope}
            data = [
                e for e in data
                if not scope_map.get(e['id']) or student_scope & scope_map[e['id']]
            ]

            exam_ids = [e['id'] for e in data]
            submissions = {
                s.exam_session_id: s
                for s in BusiExamSubmission.objects.filter(
                    exam_session_id__in=exam_ids, student=request.user,
                )
            }
            for exam in data:
                sub = submissions.get(exam['id'])
                exam['submission_status'] = sub.status if sub else None
                exam['score_obtained'] = float(sub.total_score) if sub and sub.total_score else None

        return APIResponse.success(data={'results': data, 'count': qs.count()})

    @action(methods=['get'], detail=True, url_path='my-result')
    def my_result(self, request, pk=None):
        """学生查看某场考试的成绩"""
        exam = self.get_object()
        submission = BusiExamSubmission.objects.filter(
            exam_session=exam, student=request.user,
        ).first()

        if not submission or submission.status not in ('submitted', 'auto_submitted'):
            return APIResponse.error(code=400, message='尚未提交考试')

        answers = BusiStudentAnswer.objects.filter(
            exam_session=exam, student=request.user,
        ).select_related('paper_question__question').order_by('paper_question__order')

        serializer = BusiStudentAnswerSerializer(answers, many=True)
        return APIResponse.success(data={
            'submission': BusiExamSubmissionSerializer(submission).data,
            'answers': serializer.data,
        })

    @action(methods=['get'], detail=True, url_path='grade-list')
    def grade_list(self, request, pk=None):
        """教师查看待批改列表（某场考试的所有未批改主观题）"""
        exam = self.get_object()
        ungraded = BusiStudentAnswer.objects.filter(
            exam_session=exam,
            is_correct__isnull=True,
            status='submitted',
        ).select_related('paper_question__question', 'student').order_by('student', 'paper_question__order')

        serializer = BusiStudentAnswerSerializer(ungraded, many=True)
        return APIResponse.success(data={
            'count': ungraded.count(),
            'answers': serializer.data,
        })

    @action(methods=['post'], detail=True, url_path='grade')
    def grade_answer(self, request, pk=None):
        """教师手动批改一道题"""
        answer_id = request.data.get('answer_id')
        score_obtained = request.data.get('score_obtained')

        if not answer_id or score_obtained is None:
            return APIResponse.error(code=400, message='缺少 answer_id 或 score_obtained')

        try:
            answer = BusiStudentAnswer.objects.get(
                id=answer_id, exam_session_id=pk,
            )
        except BusiStudentAnswer.DoesNotExist:
            return APIResponse.error(code=404, message='答题记录不存在')

        answer.score_obtained = score_obtained
        answer.is_correct = score_obtained > 0
        answer.status = 'graded'
        answer.save()

        # 重新计算该学生的总分
        from apps.exams.scoring import grade_exam_submission
        submission = BusiExamSubmission.objects.filter(
            exam_session_id=pk, student=answer.student,
        ).first()
        if submission:
            submission.total_score = BusiStudentAnswer.objects.filter(
                exam_session_id=pk, student=answer.student,
            ).aggregate(s=models.Sum('score_obtained'))['s'] or 0
            submission.save()

        return APIResponse.success(message='批改成功', data={
            'answer_id': answer.id,
            'score_obtained': answer.score_obtained,
        })

    def _first_error(self, errors):
        for _, msgs in errors.items():
            return str(msgs[0]) if isinstance(msgs, list) else str(msgs)
        return '参数错误'

    @action(methods=['post'], detail=False, url_path='import-candidates', permission_classes=[IsTeacher])
    def import_candidates(self, request):
        """批量导入考生（从 Excel 文件）"""
        file_obj = request.FILES.get('file')
        if not file_obj:
            return APIResponse.error(code=400, message='请上传 Excel 文件')

        try:
            import openpyxl
        except ImportError:
            return APIResponse.error(code=500, message='服务端缺少 openpyxl 依赖')

        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True)
            ws = wb.active
        except Exception as e:
            return APIResponse.error(code=400, message=f'无法解析 Excel 文件: {str(e)}')

        # 第 2 行是真正的表头：序号/村行/机构号/柜员号/姓名/岗位/分管业务/备注

        SCOPE_MAP = {
            '资产': ['资产'], '负债': ['负债'], '零售': ['零售'],
            '资产和负债': ['资产', '负债'],
            '资产、负债': ['资产', '负债'],
            '资产负债': ['资产', '负债'],
        }

        def _parse_scope(raw: str) -> list:
            """兼容解析分管业务：精确匹配 → 模糊匹配 → 空"""
            raw = raw.strip().replace(' ', '').replace('　', '')
            if not raw:
                return []
            if raw in SCOPE_MAP:
                return SCOPE_MAP[raw]
            # 模糊匹配：检查包含的关键词
            scopes = []
            if '资产' in raw:
                scopes.append('资产')
            if '负债' in raw:
                scopes.append('负债')
            if '零售' in raw:
                scopes.append('零售')
            if '制度' in raw:
                scopes.append('制度')
            return scopes

        # 第一遍：收集所有行数据到内存，按用户名合并分管业务
        total_rows = 0
        seen = {}  # username -> merged row
        for row in ws.iter_rows(min_row=3, values_only=True):
            teller_id = row[3]
            name = row[4]
            if not teller_id or not name:
                continue
            scopes = _parse_scope(str(row[6]).strip() if row[6] is not None else '')
            if not scopes:
                continue
            total_rows += 1
            uname = str(int(teller_id))
            if uname in seen:
                seen[uname]['scopes'].update(scopes)
                continue
            seen[uname] = {
                'username': uname,
                'name': str(name),
                'org_name': str(row[1]) if row[1] else '',
                'org_code': str(row[2]).strip() if row[2] is not None else '',
                'position': str(row[5]).strip() if row[5] is not None else '',
                'remark': str(row[7]).strip() if row[7] is not None else '',
                'scopes': set(scopes),
            }
        wb.close()

        # 批量操作
        from django.contrib.auth.hashers import make_password
        rows = list(seen.values())
        usernames = {r['username'] for r in rows}
        existing = {u.username: u for u in BusiUser.objects.filter(username__in=usernames)}

        to_create = []
        to_update = []
        created_count = updated_count = 0

        for r in rows:
            uname = r['username']
            defaults = {
                'first_name': r['name'], 'role': 'student',
                'org_nm': r['org_name'], 'org_no': r['org_code'],
                'position': r['position'], 'remark': r['remark'],
                'is_active': True,
            }
            if uname in existing:
                user = existing[uname]
                changed = False
                for k, v in defaults.items():
                    if getattr(user, k) != v:
                        setattr(user, k, v)
                        changed = True
                if changed:
                    to_update.append(user)
                updated_count += 1
            else:
                defaults['username'] = uname
                defaults['password'] = make_password(uname)
                defaults['password_changed'] = False
                to_create.append(BusiUser(**defaults))
                created_count += 1

        if to_create:
            BusiUser.objects.bulk_create(to_create, batch_size=200)
        if to_update:
            BusiUser.objects.bulk_update(
                to_update, ['first_name', 'org_nm', 'org_no', 'position', 'remark'],
                batch_size=200,
            )

        scope_counter = {'资产': 0, '负债': 0, 'both': 0, '零售': 0}
        for r in rows:
            combined = sorted(r['scopes'])
            BusiUser.objects.filter(username=r['username']).update(business_scope=combined)
            key = 'both' if len(combined) > 1 else combined[0] if combined else ''
            if key:
                scope_counter[key] = scope_counter.get(key, 0) + 1

        return APIResponse.success(data={
            'created': created_count,
            'updated': updated_count,
            'total': len(rows),
            'total_rows': total_rows,
            'by_scope': scope_counter,
        }, message=f'导入完成：共 {total_rows} 行数据，{len(rows)} 名考生（新增 {created_count}，更新 {updated_count}）')

    # ==================== P3: 成绩管理 ====================

    @action(methods=['get'], detail=True, url_path='results', permission_classes=[IsTeacher])
    def exam_results(self, request, pk=None):
        """查看某场考试的全部考生成绩（按业务分榜）"""
        exam = self.get_object()
        submissions = BusiExamSubmission.objects.filter(
            exam_session=exam, is_deleted=False,
        ).select_related('student', 'adjusted_by').order_by('-total_score')

        results = []
        for sub in submissions:
            adj = sub.adjusted_by
            results.append({
                'id': sub.id,
                'student_id': sub.student_id,
                'student_username': sub.student.username,
                'student_name': sub.student.first_name or sub.student.username,
                'org_no': getattr(sub.student, 'org_no', ''),
                'org_nm': getattr(sub.student, 'org_nm', ''),
                'business_scope': getattr(sub.student, 'business_scope', []),
                'status': sub.status,
                'total_score': float(sub.total_score) if sub.total_score else 0,
                'float_score': float(sub.float_score) if sub.float_score else 0,
                'float_score_reason': sub.float_score_reason or '',
                'final_score': (float(sub.total_score or 0) + float(sub.float_score or 0)),
                'submit_time': sub.submit_time.strftime('%Y-%m-%d %H:%M:%S') if sub.submit_time else '',
                'adjusted_by': f'{adj.username}-{adj.first_name}' if adj else '',
                'adjusted_at': sub.adjusted_at.strftime('%Y-%m-%d %H:%M:%S') if sub.adjusted_at else '',
            })

        # 分榜：考试未限定范围时按考生业务范围分，否则双重过滤
        exam_scope = getattr(exam, 'exam_scope', []) or []
        if not exam_scope:
            asset_list = [r for r in results if isinstance(r['business_scope'], list) and '资产' in r['business_scope']]
            liability_list = [r for r in results if isinstance(r['business_scope'], list) and '负债' in r['business_scope']]
        else:
            asset_list = [r for r in results
                          if '资产' in exam_scope and isinstance(r['business_scope'], list) and '资产' in r['business_scope']]
            liability_list = [r for r in results
                              if '负债' in exam_scope and isinstance(r['business_scope'], list) and '负债' in r['business_scope']]
        asset_list.sort(key=lambda x: x['final_score'], reverse=True)
        liability_list.sort(key=lambda x: x['final_score'], reverse=True)

        return APIResponse.success(data={
            'all': results,
            'asset_ranking': asset_list,
            'liability_ranking': liability_list,
            'exam_name': exam.name,
            'float_enabled': hasattr(exam, 'config') and exam.config.float_score_enabled,
        })

    @action(methods=['post'], detail=True, url_path='adjust-score', permission_classes=[IsTeacher])
    def adjust_score(self, request, pk=None):
        """评委调整某考生浮动分"""
        submission_id = request.data.get('submission_id')
        float_score = request.data.get('float_score')
        reason = request.data.get('reason', '')

        if submission_id is None or float_score is None:
            return APIResponse.error(code=400, message='请提供 submission_id 和 float_score')

        try:
            sub = BusiExamSubmission.objects.get(id=submission_id, exam_session_id=pk)
        except BusiExamSubmission.DoesNotExist:
            return APIResponse.error(code=404, message='提交记录不存在')

        # 非管理员不能重复修改已打过的浮动分
        if sub.float_score is not None and not request.user.is_admin:
            return APIResponse.error(code=403, message='浮动分已打过，仅管理员可多次修改，请联系管理员调整')

        try:
            fs = float(float_score)
        except (ValueError, TypeError):
            return APIResponse.error(code=400, message='浮动分必须为数字')

        # 范围校验
        config = getattr(sub.exam_session, 'config', None)
        if config and config.float_score_enabled:
            if fs < config.float_score_min or fs > config.float_score_max:
                return APIResponse.error(code=400,
                    message=f'浮动分超出范围 [{config.float_score_min}, {config.float_score_max}]')

        sub.float_score = fs
        sub.float_score_reason = str(reason)[:500] if reason else ''
        sub.adjusted_by = request.user
        sub.adjusted_at = timezone.now()
        sub.save()

        return APIResponse.success(data={
            'id': sub.id,
            'student_name': sub.student.first_name or sub.student.username,
            'total_score': float(sub.total_score or 0),
            'float_score': float(sub.float_score or 0),
            'final_score': float(sub.total_score or 0) + float(sub.float_score or 0),
        }, message='浮动分调整成功')

    @action(methods=['get'], detail=True, url_path='export', permission_classes=[IsTeacher])
    def export_results(self, request, pk=None):
        """导出成绩为 Excel"""
        exam = self.get_object()
        submissions = BusiExamSubmission.objects.filter(
            exam_session=exam, is_deleted=False,
        ).select_related('student', 'adjusted_by')

        import io
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse

        wb = openpyxl.Workbook()

        for scope, label in [('资产', '资产榜'), ('负债', '负债榜')]:
            if scope == 'asset':
                sheet = wb.active
                sheet.title = label
            else:
                sheet = wb.create_sheet(label)

            headers = ['排名', '姓名', '村行', '基础分', '浮动分', '最终得分', '分管业务', '提交时间']
            for c, h in enumerate(headers, 1):
                cell = sheet.cell(1, c, h)
                cell.font = Font(bold=True)

            rows = []
            for sub in submissions:
                bs = getattr(sub.student, 'business_scope', [])
                if not isinstance(bs, list) or scope not in bs:
                    continue
                rows.append({
                    'name': sub.student.first_name or sub.student.username,
                    'org': getattr(sub.student, 'org_nm', ''),
                    'total': float(sub.total_score or 0),
                    'float': float(sub.float_score or 0),
                    'final': float(sub.total_score or 0) + float(sub.float_score or 0),
                    'scope': ', '.join(bs),
                    'time': sub.submit_time.strftime('%Y-%m-%d %H:%M') if sub.submit_time else '',
                })
            rows.sort(key=lambda x: x['final'], reverse=True)

            for i, r in enumerate(rows, 2):
                sheet.cell(i, 1, i - 1).alignment = Alignment(horizontal='center')
                sheet.cell(i, 2, r['name'])
                sheet.cell(i, 3, r['org'])
                sheet.cell(i, 4, r['total']).alignment = Alignment(horizontal='center')
                sheet.cell(i, 5, r['float']).alignment = Alignment(horizontal='center')
                sheet.cell(i, 6, r['final']).alignment = Alignment(horizontal='center')
                sheet.cell(i, 7, r['scope'])
                sheet.cell(i, 8, r['time'])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{exam.name}_成绩.xlsx"'
        return response
